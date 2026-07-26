#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
지역부 회의자료(상반기 Review & 하반기 영업전략) → 엑셀 변환기

HWP → PDF 로 만든 '26년 상반기 Review & 하반기 수원지역부 영업전략' 보고서를
편집 가능한 엑셀(.xlsx)로 변환합니다.

부장님 피드백 반영:
  - 각 팀별 하반기 목표는 좋으나, "7월에 몇 개 했는지" 체크가 필요.
  - 각 항목별 "7월 전개 내용"을 기입할 수 있도록 컬럼 추가.
→ '★7월 팀별 전개체크' 시트에 팀별·항목별 7월 실적/전개 내용 입력란을 제공하고,
   달성률은 수식으로 자동 계산되도록 구성했습니다.

사용법:
    python regional_meeting_to_excel.py [출력파일.xlsx]
    (인자를 생략하면 '수원지역부_회의자료_2026H2.xlsx' 로 저장)

필요 패키지:
    pip install openpyxl
"""
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 공통 스타일
# ---------------------------------------------------------------------------
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D9E1F2"
LIGHT2 = "EAF0FA"
YELLOW = "FFF2CC"
GREEN = "E2EFDA"
GREY = "F2F2F2"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TITLE_FONT = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
H1_FONT = Font(name="맑은 고딕", size=13, bold=True, color="FFFFFF")
H2_FONT = Font(name="맑은 고딕", size=11, bold=True, color=NAVY)
HEAD_FONT = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="맑은 고딕", size=10, bold=True)
BODY = Font(name="맑은 고딕", size=10)
NOTE = Font(name="맑은 고딕", size=9, italic=True, color="808080")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

TEAMS = [
    "수원영업1팀", "수원영업2팀", "수원영업3팀", "수원영업4팀", "수원영업5팀",
    "수원영업6팀", "수원영업7팀", "수원영업8팀", "수원영업9팀",
]


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def put(ws, cell, value, font=BODY, align=CENTER, fillc=None, border=True, numfmt=None):
    c = ws[cell]
    c.value = value
    c.font = font
    c.alignment = align
    if fillc:
        c.fill = fill(fillc)
    if border:
        c.border = BORDER
    if numfmt:
        c.number_format = numfmt
    return c


def banner(ws, row, ncols, text, font=H2_FONT, fillc=LIGHT):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font
    c.alignment = LEFT if font is H2_FONT else CENTER
    c.fill = fill(fillc)
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).border = BORDER
    ws.row_dimensions[row].height = 22 if font is H2_FONT else 18


def month_table(ws, start_row, unit, title, y25, y26, yoy, rank, ncols_label="지역부"):
    """25/26년 월별 추이 6개월 표를 그린다. 반환: 다음 빈 행."""
    r = start_row
    months25 = ["25년 1월", "25년 2월", "25년 3월", "25년 4월", "25년 5월", "25년 6월"]
    months26 = ["26년 1월", "26년 2월", "26년 3월", "26년 4월", "26년 5월", "26년 6월"]

    banner(ws, r, 7, title, font=H2_FONT, fillc=LIGHT)
    r += 1
    put(ws, f"A{r}", unit, NOTE, RIGHT, border=False)
    ws.merge_cells(f"A{r}:G{r}")
    r += 1

    put(ws, f"A{r}", ncols_label, HEAD_FONT, CENTER, BLUE)
    for i, m in enumerate(months25):
        put(ws, f"{get_column_letter(2 + i)}{r}", m, HEAD_FONT, CENTER, BLUE)
    r += 1
    put(ws, f"A{r}", "25년 실적", BOLD, CENTER, GREY)
    for i, v in enumerate(y25):
        put(ws, f"{get_column_letter(2 + i)}{r}", v, BODY)
    r += 1
    put(ws, f"A{r}", "수원지역부", HEAD_FONT, CENTER, BLUE)
    for i, m in enumerate(months26):
        put(ws, f"{get_column_letter(2 + i)}{r}", m, HEAD_FONT, CENTER, BLUE)
    r += 1
    put(ws, f"A{r}", "26년 실적", BOLD, CENTER, GREY)
    for i, v in enumerate(y26):
        put(ws, f"{get_column_letter(2 + i)}{r}", v, BODY)
    r += 1
    put(ws, f"A{r}", "전년비", BOLD, CENTER, YELLOW)
    for i, v in enumerate(yoy):
        put(ws, f"{get_column_letter(2 + i)}{r}", v, BODY)
    r += 1
    put(ws, f"A{r}", "전사 순위", BOLD, CENTER, GREEN)
    for i, v in enumerate(rank):
        put(ws, f"{get_column_letter(2 + i)}{r}", v, BODY)
    r += 1
    return r + 1


def bullets(ws, start_row, ncols, lines):
    r = start_row
    for line in lines:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
        c = ws.cell(row=r, column=1, value=line)
        c.font = BODY
        c.alignment = LEFT
        ws.row_dimensions[r].height = 16
        r += 1
    return r + 1


# ---------------------------------------------------------------------------
# 표지
# ---------------------------------------------------------------------------
def build_cover(wb):
    ws = wb.active
    ws.title = "표지"
    set_widths(ws, [3, 22, 22, 22, 22, 22, 3])
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:F2")
    put(ws, "B2", "대 외 비", Font(name="맑은 고딕", size=11, bold=True, color="C00000"),
        RIGHT, border=False)

    ws.merge_cells("B4:F6")
    put(ws, "B4", "26년 상반기 Review & 하반기\n수원지역부 영업전략", TITLE_FONT, CENTER, NAVY, border=False)
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 30

    ws.merge_cells("B8:F8")
    put(ws, "B8", "2권역  수원지역부", H2_FONT, CENTER, border=False)
    ws.merge_cells("B9:F9")
    put(ws, "B9", "보고일 : 2026.07.13", BODY, CENTER, border=False)

    ws.merge_cells("B12:F12")
    put(ws, "B12", "목  차", Font(name="맑은 고딕", size=12, bold=True, color=NAVY), LEFT, border=False)

    toc = [
        "1. 일반상품 재고 및 점두 진열",
        "2. 상권최적화 추진",
        "3. 간편식사 확대 방안",
        "4. GET커피 잔수·매출 확대",
        "5. 즉석조리 가동률·매출 극대화",
        "★ 7월 팀별 전개체크  (부장님 피드백 반영 : 팀별 7월 실적/전개 내용 기입)",
    ]
    r = 13
    for t in toc:
        ws.merge_cells(f"B{r}:F{r}")
        put(ws, f"B{r}", t, BODY if not t.startswith("★") else BOLD, LEFT, border=False)
        r += 1

    r += 1
    ws.merge_cells(f"B{r}:F{r+4}")
    note = ("[안내] 이 엑셀은 HWP→PDF 보고서를 편집 가능한 엑셀로 변환한 자료입니다.\n"
            "· 각 전략 시트 : 원본 보고서의 실적표를 그대로 옮겨 편집 가능하게 정리\n"
            "· ★7월 팀별 전개체크 시트 : 팀별 하반기 목표 대비 7월 실적/전개 내용 (달성률 자동 계산)\n"
            "· 연장진열대 : 취합 리스트 설치일자 기준 재집계 → 상반기(~6월) 75점 + 7월 신규 4점으로 분리\n"
            "· 점두진열 : 진열점입력(7월 분석) 기준 팀별 진열점 현황 반영 (노란색 칸이 입력란)")
    put(ws, f"B{r}", note, Font(name="맑은 고딕", size=9, color="595959"), LEFT, GREY, border=False)


# ---------------------------------------------------------------------------
# 1. 일반상품 재고 및 점두 진열
# ---------------------------------------------------------------------------
def build_s1(wb):
    ws = wb.create_sheet("1.일반상품재고")
    set_widths(ws, [16, 12, 12, 12, 12, 12, 12])
    banner(ws, 1, 7, "1. 일반상품 재고 및 점두 진열", font=H1_FONT, fillc=NAVY)
    ws.row_dimensions[1].height = 26

    r = 3
    r = bullets(ws, r, 7, [
        "가. 재고 목표 : 25년 7월 19,945천원 → 목표 20,942천원(105%) / 26년 7월 1주차 20,788천원",
    ])
    r = month_table(
        ws, r, "(단위 : 천원 / %)", "1) 26년도 일반상품 재고 실적 추이",
        [19624, 19909, 19305, 19954, 19631, 19784],
        [20372, 20437, 20050, 19812, 19907, 20655],
        ["100.8%", "99.7%", "103.9%", "99.3%(물류파업)", "101.4%", "104.4%"],
        ["2위", "2위", "2위", "11위", "2위", "1위"],
    )

    # 연장진열대 팀표
    banner(ws, r, 7, "나. 관련 전략  ―  1) 연장진열대 및 점두진열 전개", font=H2_FONT, fillc=LIGHT)
    r += 1
    # ※ 연장진열대 : 취합 리스트 설치일자 기준 재집계 (상반기=~6월 / 7월 분리)
    #   기존 '상반기 79점' 중 4점이 실제로는 7월 설치 → ~6월 75점 + 7월 4점
    r = team_progress_block(
        ws, r, "연장진열대 (단위: 점)  ※ 설치일자 기준 재집계 : 상반기=~6월 / 7월 분리",
        headers=["구분", "~6월 완료", "7월 신규", "하반기 목표", "합계"],
        region=[75, 4, 26, 105],
        rows={
            "수원영업1팀": [14, 0, 3, 17], "수원영업2팀": [8, 1, 7, 16], "수원영업3팀": [3, 1, 1, 5],
            "수원영업4팀": [6, 0, 3, 9], "수원영업5팀": [5, 1, 1, 7], "수원영업6팀": [9, 0, 0, 9],
            "수원영업7팀": [11, 0, 3, 14], "수원영업8팀": [11, 0, 3, 14], "수원영업9팀": [8, 1, 5, 14],
        },
    )
    r = bullets(ws, r, 7, [
        "  - 연장진열대 ~6월 75점 완료 + 7월 신규 4점(누계 79점), 하반기 목표 26점 진행",
        "     · 7월 신규 : 권선아이파크점(7/2), 화성봉황점(7/6), 산본중앙로점(7/6), 행정드림점(7/30·예정)",
        "  - 연장진열대 79점 기준 6월 30일 전년 비교시 일반상품 재고 +939천원, 전년비 105.3% 증가",
        "     ※ 연장진열대 설치점 일매출 전년비 109.6%, 150천원 증가",
        "  - 하반기 26점 확대 목표 전개, 리뉴얼 진행시 일체형 진열대 높이 조정 추진",
    ])

    # ※ 점두진열 : 7월 진열점 현황(Material B '진열점입력') = 105점 추가 표기
    r = team_progress_block(
        ws, r, "점두진열 (단위: 점)  ※ 7월 진열점 현황 = 진열점입력(7월 분석) 기준",
        headers=["구분", "상반기 완료", "하반기 목표", "합계", "7월 진열점(현황)"],
        region=[108, 45, 153, 105],
        rows={
            "수원영업1팀": [12, 2, 14, 12], "수원영업2팀": [20, 3, 23, 18], "수원영업3팀": [6, 2, 8, 6],
            "수원영업4팀": [13, 8, 21, 13], "수원영업5팀": [14, 12, 26, 14], "수원영업6팀": [9, 11, 20, 9],
            "수원영업7팀": [16, 2, 18, 16], "수원영업8팀": [6, 4, 10, 6], "수원영업9팀": [12, 1, 13, 11],
        },
    )
    r = bullets(ws, r, 7, [
        "  - 6월 권역(44점포) + 지역부(64점포) 총 108점 점두 진열 진행",
        "  - 지역부 진열 7월 기준 64점 중 52점(81.2%) 일매출 향상 효과 有 (*일부점포 데이터 추출 오류)",
        "  - 7월 진열점 현황 : '진열점입력'(7월 분석) 기준 105점 (날짜 필드 없어 7월 신규분 분리 불가)",
    ])

    # APPENDIX 1
    banner(ws, r, 7, "※ APPENDIX 1. 7월 팀별 일매출 대표 상승점 (단위: 원)", font=H2_FONT, fillc=LIGHT2)
    r += 1
    heads = ["영업팀", "점포명", "상품", "26.07 일매출", "25.07 일매출", "신장률"]
    # 6열 표 (A~F)
    for i, h in enumerate(heads):
        put(ws, f"{get_column_letter(1 + i)}{r}", h, HEAD_FONT, CENTER, BLUE)
    r += 1
    app1 = [
        ["7팀", "동탄하이페리온점", "신라면로제봉지, 배홍동비빔면 외 2종", 14423, 605, "2,384.7%"],
        ["5팀", "산본하이어스점", "배홍동비빔면, 하이네켄캔 외 2종", 45272, 4576, "989.3%"],
        ["1팀", "광교중흥점", "신라면, 짜파게티", 4973, 987, "503.8%"],
        ["4팀", "장안빌리지점", "신라면, 카스캔473ml6입번들팩", 11142, 2239, "497.7%"],
        ["9팀", "화성상신중학교점", "신라면, 팔도비빔면", 19477, 4063, "479.4%"],
    ]
    for row in app1:
        for i, v in enumerate(row):
            nf = "#,##0" if i in (3, 4) else None
            al = LEFT if i in (1, 2) else CENTER
            put(ws, f"{get_column_letter(1 + i)}{r}", v, BODY, al, numfmt=nf)
        r += 1
    r += 1

    # APPENDIX 2
    banner(ws, r, 7, "※ APPENDIX 2. 6월 팀별 점두전개점 일매출 신장률 (단위: 원)", font=H2_FONT, fillc=LIGHT2)
    r += 1
    heads = ["영업팀", "점포수", "26.06 일매출", "25.06 일매출", "신장률"]
    for i, h in enumerate(heads):
        put(ws, f"{get_column_letter(1 + i)}{r}", h, HEAD_FONT, CENTER, BLUE)
    r += 1
    app2 = [
        ["수원영업1팀", 12, 2203824, 1928774, "114.26%"],
        ["수원영업2팀", 20, 2600600, 2327588, "111.73%"],
        ["수원영업3팀", 6, 2046884, 1782673, "114.82%"],
        ["수원영업4팀", 13, 2169235, 2027682, "106.98%"],
        ["수원영업5팀", 14, 2485621, 2411177, "103.09%"],
        ["수원영업6팀", 9, 2493883, 2340683, "106.55%"],
        ["수원영업7팀", 16, 2526522, 2144650, "117.81%"],
        ["수원영업8팀", 6, 2502098, 2456324, "101.86%"],
        ["수원영업9팀", 11, 2084906, 2126290, "98.05%"],
    ]
    for row in app2:
        for i, v in enumerate(row):
            nf = "#,##0" if i in (2, 3) else None
            al = LEFT if i == 0 else CENTER
            put(ws, f"{get_column_letter(1 + i)}{r}", v, BODY, al, numfmt=nf)
        r += 1
    r = bullets(ws, r + 1, 7, ["  - 수원8팀 화성파인밸리점 6/30 개점에 따른 전년비 왜곡 보정"])


def team_progress_block(ws, start_row, title, headers, region, rows):
    """지역부 + 9팀 실적표 (완료/목표/합계 형태). 반환: 다음 행."""
    r = start_row
    ncol = len(headers)
    banner(ws, r, 7, title, font=H2_FONT, fillc=LIGHT2)
    r += 1
    for i, h in enumerate(headers):
        put(ws, f"{get_column_letter(1 + i)}{r}", h, HEAD_FONT, CENTER, BLUE)
    r += 1
    put(ws, f"A{r}", "수원지역부", BOLD, CENTER, YELLOW)
    for i, v in enumerate(region):
        put(ws, f"{get_column_letter(2 + i)}{r}", v, BOLD, CENTER, YELLOW)
    r += 1
    for team, vals in rows.items():
        put(ws, f"A{r}", team, BODY, CENTER, GREY)
        for i, v in enumerate(vals):
            put(ws, f"{get_column_letter(2 + i)}{r}", v, BODY)
        r += 1
    return r + 1


# ---------------------------------------------------------------------------
# 2. 상권최적화
# ---------------------------------------------------------------------------
def build_s2(wb):
    ws = wb.create_sheet("2.상권최적화")
    set_widths(ws, [18, 14, 14, 14, 14, 12, 12])
    banner(ws, 1, 7, "2. 상권최적화 추진", font=H1_FONT, fillc=NAVY)
    ws.row_dimensions[1].height = 26

    r = 3
    banner(ws, r, 4, "1) 상권최적화 실적 현황 (6월 마감) (단위: 점)", font=H2_FONT, fillc=LIGHT2)
    r += 1
    for i, h in enumerate(["상권최적화", "6월 누계", "하반기 계획", "합계"]):
        put(ws, f"{get_column_letter(1 + i)}{r}", h, HEAD_FONT, CENTER, BLUE)
    r += 1
    for row in [["이전/대체", 12, 9, 22], ["통합", 2, 21, 23], ["확장", 1, 14, 15]]:
        for i, v in enumerate(row):
            put(ws, f"{get_column_letter(1 + i)}{r}", v, BODY, CENTER,
                GREY if i == 0 else None)
        r += 1
    for i, v in enumerate(["합계", 15, 44, 60]):
        put(ws, f"{get_column_letter(1 + i)}{r}", v, BOLD, CENTER, YELLOW)
    r += 2

    banner(ws, r, 5, "2) 상권최적화 성공 사례 (단위: 천원, %)", font=H2_FONT, fillc=LIGHT2)
    r += 1
    for i, h in enumerate(["점포명", "최적화 유형", "최적화 익월 일매출", "전년 동월 일매출", "신장률"]):
        put(ws, f"{get_column_letter(1 + i)}{r}", h, HEAD_FONT, CENTER, BLUE)
    r += 1
    cases = [
        ["탑동중앙점", "이전/대체", 4783, 1206, "397%"],
        ["화성대림점", "이전/대체", 2839, 1421, "200%"],
        ["수원안다미로점", "이전/대체", 4405, 1675, "263%"],
        ["광교웰빙타운점", "통합", 2650, 2226, "119%"],
        ["화성병점원룸점", "통합", 1850, 1392, "133%"],
        ["봉담에듀포레점", "확장", 1782, 1191, "150%"],
    ]
    for row in cases:
        for i, v in enumerate(row):
            nf = "#,##0" if i in (2, 3) else None
            al = LEFT if i == 0 else CENTER
            put(ws, f"{get_column_letter(1 + i)}{r}", v, BODY, al, numfmt=nf)
        r += 1

    r = bullets(ws, r + 1, 7, [
        "  - 상권최적화 진행점 이전/대체 12점, 통합 2점, 확장 1건 6월 마무리",
        "  - 목표대비 +1점으로 권역 최상위 실적 진행",
        "",
        "3) 하반기 상권최적화 진행 방안",
        "  - 주차별 유사업종A 방문, 후보점 발굴 진행 (7월 주차별 취합 진행)",
        "  - 스마트그로서리 기반 유사업종B 후보점 발굴 추가 진행",
        "  - 지역별 SP + 영업팀장 동행 순회를 통한 상권최적화 인사이트 확대, 협업 시너지 창출",
    ])


# ---------------------------------------------------------------------------
# 3. 간편식사
# ---------------------------------------------------------------------------
def build_s3(wb):
    ws = wb.create_sheet("3.간편식사")
    set_widths(ws, [16, 12, 12, 12, 12, 12, 12])
    banner(ws, 1, 7, "3. 간편식사 확대 방안", font=H1_FONT, fillc=NAVY)
    ws.row_dimensions[1].height = 26

    r = 3
    r = month_table(
        ws, r, "(단위 : 일매출 / %)", "1) 간편식사 실적 추이",
        [78300, 76335, 89562, 95721, 92935, 96908],
        [86287, 82746, 97623, 101274, 99626, 104075],
        ["110.2%", "108.4%", "109.0%", "105.8%(물류파업)", "107.2%", "104.4%"],
        ["3위", "4위", "11위", "13위", "18위", "5위"],
    )
    r = bullets(ws, r, 7, [
        "2) 목표",
        "  - 발주 : 25년 7월 40.5 → 목표 42.5 (105%) / 26년 7월 1주차 현재 40.5",
        "  - 일매출 : 25년 7월 96,044원 → 목표 100,846원 (105%) / 26년 7월 1주차 현재 97,965원",
    ])

    banner(ws, r, 7, "나. 간편식사 전개 방안 (하반기 진행점, 단위: 점)", font=H2_FONT, fillc=LIGHT)
    r += 1
    # 세 항목의 하반기 진행점 팀표 (진행점만 존재)
    r = team_target_block(
        ws, r, "1) 주먹밥 + 음료 (학생층 타겟팅)",
        region=87,
        rows={"수원영업1팀": 14, "수원영업2팀": 13, "수원영업3팀": 9, "수원영업4팀": 10,
              "수원영업5팀": 12, "수원영업6팀": 13, "수원영업7팀": 6, "수원영업8팀": 1, "수원영업9팀": 9},
    )
    r = bullets(ws, r, 7, [
        "  - 하교 후 14~16시, 학원이동 19~21시 삼각김밥+음료 증정행사 진행",
        "  - 삼각김밥 증정행사 및 모음진열·POP 확대 87점 진행",
        "  - 삼각김밥 소분류 일매출 38,103원, 6월 기준 전년비 105.4% 달성 (6월 52점 기준)",
    ])
    r = team_target_block(
        ws, r, "2) 헬시플레저 샌드위치 SKU 최적화",
        region=97,
        rows={"수원영업1팀": 14, "수원영업2팀": 18, "수원영업3팀": 6, "수원영업4팀": 10,
              "수원영업5팀": 12, "수원영업6팀": 10, "수원영업7팀": 6, "수원영업8팀": 9, "수원영업9팀": 12},
    )
    r = bullets(ws, r, 7, [
        "  - 저당 샌드위치 + 지중해 올리브 컨셉 도입, 진열 정비 총 97점 진행",
        "  - 일매출 31,940원, 전년비 110.6%, 전월비 103.2% 기록 (6월 58점 기준)",
    ])
    r = team_target_block(
        ws, r, "3) 간편식사 아크릴 진열대 도입",
        region=147,
        rows={"수원영업1팀": 18, "수원영업2팀": 22, "수원영업3팀": 25, "수원영업4팀": 14,
              "수원영업5팀": 13, "수원영업6팀": 13, "수원영업7팀": 13, "수원영업8팀": 16, "수원영업9팀": 13},
    )
    r = bullets(ws, r, 7, [
        "  - ① 발주 확대형 보조진열대, ② 계단식 보조진열대 TWO-트랙 전개",
        "  - 발주 확대형 72점(144개), 계단식 75점(150개) 도입 예정 (예산 전용중)",
    ])


def team_target_block(ws, start_row, title, region, rows):
    """하반기 진행점(목표)만 존재하는 팀표. 반환: 다음 행."""
    r = start_row
    banner(ws, r, 3, title, font=H2_FONT, fillc=LIGHT2)
    r += 1
    for i, h in enumerate(["구분", "하반기 진행점(목표)"]):
        put(ws, f"{get_column_letter(1 + i)}{r}", h, HEAD_FONT, CENTER, BLUE)
    r += 1
    put(ws, f"A{r}", "수원지역부", BOLD, CENTER, YELLOW)
    put(ws, f"B{r}", region, BOLD, CENTER, YELLOW)
    r += 1
    for team, v in rows.items():
        put(ws, f"A{r}", team, BODY, CENTER, GREY)
        put(ws, f"B{r}", v, BODY)
        r += 1
    return r + 1


# ---------------------------------------------------------------------------
# 4. GET커피
# ---------------------------------------------------------------------------
def build_s4(wb):
    ws = wb.create_sheet("4.GET커피")
    set_widths(ws, [16, 12, 12, 12, 12, 12, 12])
    banner(ws, 1, 7, "4. GET커피 잔수·매출 확대", font=H1_FONT, fillc=NAVY)
    ws.row_dimensions[1].height = 26

    r = 3
    r = bullets(ws, r, 7, [
        "가. 목표",
        "  - 잔수 : 25년 7월 6.3잔 → 26년 7월 현재 7.1잔 (9잔 목표)",
        "  - 일매출 : 25년 7월 4,785원 → 26년 7월 현재 5,943원 (7,000원 목표)",
    ])
    r = month_table(
        ws, r, "(단위 : 잔 / %)", "1) GET커피 실적 추이",
        [5.4, 5.7, 5.9, 5.9, 6.1, 6.0],
        [6.6, 6.7, 7.3, 7.6, 7.2, 7.5],
        ["123.1%", "116.7%", "122.3%", "128.3%(물류파업)", "118.7%", "123.7%"],
        ["15위", "15위", "14위", "14위", "14위", "14위"],
    )

    banner(ws, r, 7, "나. 라심발리 도입", font=H2_FONT, fillc=LIGHT)
    r += 1
    r = bullets(ws, r, 7, [
        "  - 수원지역부 라심발리 도입점 전체 73%(총 882점 중 644점) / 권역 65.5%",
        "  - 라심발리 도입점 커피(자동) 일매출 전년 대비 6월 8,612원(178% 상승)",
    ])
    r = team_progress_block(
        ws, r, "라심발리 교체 (단위: 점)",
        headers=["구분", "완료점(상반기)", "하반기 교체", "합계"],
        region=[59, 12, 71],
        rows={
            "수원영업1팀": [5, 1, 6], "수원영업2팀": [8, 1, 9], "수원영업3팀": [6, 4, 10],
            "수원영업4팀": [5, 0, 5], "수원영업5팀": [2, 0, 2], "수원영업6팀": [5, 1, 6],
            "수원영업7팀": [10, 2, 12], "수원영업8팀": [10, 2, 12], "수원영업9팀": [8, 1, 9],
        },
    )

    banner(ws, r, 7, "다. 프라페 도입 및 운영", font=H2_FONT, fillc=LIGHT)
    r += 1
    r = bullets(ws, r, 7, [
        "  - 7월 수원지역부 2종 운영률 58%, 권역 내 1등",
        "  - 5,6월 로드사이드(81점) 판매 1,028개, 산업지대(30점) 408개 등 외곽지 높은 판매량",
        "  2) 외곽지 집중 전개 (수원8, 9팀) (단위: 개 / %)",
    ])
    heads = ["영업팀", "25.06 수량", "26.06 수량", "25.07 수량", "~7.12 수량", "6월 전년비", "7월 전년비"]
    for i, h in enumerate(heads):
        put(ws, f"{get_column_letter(1 + i)}{r}", h, HEAD_FONT, CENTER, BLUE)
    r += 1
    for row in [["수원영업8팀", 257, 740, 169, 189, "288%", "112%"],
                ["수원영업9팀", 382, 547, 257, 204, "143%", "79%"]]:
        for i, v in enumerate(row):
            al = LEFT if i == 0 else CENTER
            put(ws, f"{get_column_letter(1 + i)}{r}", v, BODY, al,
                numfmt="#,##0" if 1 <= i <= 4 else None)
        r += 1
    r = bullets(ws, r + 1, 7, [
        "  - 프라페 모형 POP, 냉동R/I 진열, 라심발리 측면 POP로 홍보물 노출 극대화",
        "  - 로드사이드 점두 단글자 홍보물 전개, 차량 이용 고객 시인성 확보로 구매 유입 성공",
        "  - 외곽지 점포 우선 전개, 특수입지(구내공장·산업지대) 추가 전개",
    ])


# ---------------------------------------------------------------------------
# 5. 즉석조리
# ---------------------------------------------------------------------------
def build_s5(wb):
    ws = wb.create_sheet("5.즉석조리")
    set_widths(ws, [16, 12, 12, 12, 12, 12, 12])
    banner(ws, 1, 7, "5. 즉석조리 가동률·매출 극대화", font=H1_FONT, fillc=NAVY)
    ws.row_dimensions[1].height = 26

    r = 3
    r = month_table(
        ws, r, "(단위 : %)", "1) 즉석조리 가동률 추이 (후라이드)",
        ["58.5%", "60.0%", "66.0%", "69.4%", "67.5%", "64.7%"],
        ["62.1%", "59.1%", "65.7%", "68.9%", "62.0%", "65.0%"],
        ["106.1%", "98.5%", "99.6%", "99.3%(물류파업)", "91.9%", "100.4%"],
        ["22위", "22위", "22위", "19위", "22위", "22위"],
    )
    r = month_table(
        ws, r, "(단위 : %)", "2) 즉석조리 가동률 추이 (고구마)",
        ["70.6%", "66.7%", "56.1%", "40.4%", "38.7%", "28.6%"],
        ["78.7%", "71.5%", "60.2%", "52.3%", "49.7%", "40.5%"],
        ["111.4%", "107.2%", "107.3%", "129.6%(물류파업)", "128.4%", "123.7%"],
        ["15위", "15위", "14위", "14위", "17위", "20위"],
    )

    banner(ws, r, 7, "가. 운영 현황 및 목표 (07.12) ― 가동률 20% 미만점 현황 (단위: 점)", font=H2_FONT, fillc=LIGHT)
    r += 1
    r = bullets(ws, r, 7, [
        "  - 즉석 후라이드 가동률 : 당월 65.1% / 권역 68.9% / 전사 71.1%",
        "  - 즉석 고구마 가동률 : 당월 44.3% / 권역 54.8% / 전사 54.3%",
    ])
    for i, h in enumerate(["구분", "후라이드 20%미만", "고구마 20%미만"]):
        put(ws, f"{get_column_letter(1 + i)}{r}", h, HEAD_FONT, CENTER, BLUE)
    r += 1
    fry = {"수원영업1팀": 7, "수원영업2팀": 8, "수원영업3팀": 2, "수원영업4팀": 6, "수원영업5팀": 4,
           "수원영업6팀": 5, "수원영업7팀": 1, "수원영업8팀": 1, "수원영업9팀": 4}
    goguma = {"수원영업1팀": 15, "수원영업2팀": 21, "수원영업3팀": 4, "수원영업4팀": 13, "수원영업5팀": 15,
              "수원영업6팀": 7, "수원영업7팀": 8, "수원영업8팀": 2, "수원영업9팀": 8}
    put(ws, f"A{r}", "수원지역부", BOLD, CENTER, YELLOW)
    put(ws, f"B{r}", 38, BOLD, CENTER, YELLOW)
    put(ws, f"C{r}", 93, BOLD, CENTER, YELLOW)
    r += 1
    for team in TEAMS:
        put(ws, f"A{r}", team, BODY, CENTER, GREY)
        put(ws, f"B{r}", fry[team], BODY)
        put(ws, f"C{r}", goguma[team], BODY)
        r += 1
    r += 1

    r = bullets(ws, r, 7, [
        "나. 가동률 저조 원인 분석",
        "  - 미가동 권역 내 최저[고구마 63점, 후라이드 11점], 실질 가동률 차이 多",
        "  - 현장지도시 매출 타각, 단기간 운영 등이 가동률 부진 주요 원인",
        "  - 실질 가동률 극대화를 위한 20% 미만점 개선 추진",
        "",
        "다. 주차별 팀별 미가동점 점검",
        "  - 재가동 협의 불가시 마스터 삭제, 고구마 기기·튀김기 철수 진행",
        "  - 철수 거부시 ① 점간이동 ② 마스터 삭제 ③ 집기 원복 차선책 검토",
        "",
        "라. 가동·매출 극대화 방안 (9팀 사례 : 해동형 닭강정 비조리 상시 진열)",
        "  - 기아복지관점 6월 111개 → 7월 12일 누계 72개",
        "  - QC팀·국민신문고 확인 : 영양성분·원산지 표시 적합시 판매 가능",
        "  - 7/20 9팀 사례발표 공유 → 각 팀별 운영 추진",
    ])


# ---------------------------------------------------------------------------
# ★ 7월 팀별 전개체크  (부장님 피드백 반영)
# ---------------------------------------------------------------------------
def build_check(wb):
    ws = wb.create_sheet("★7월 팀별 전개체크")
    # A항목 B팀 C상반기완료 D하반기목표 E7월완료(입력) F7월누계 G달성률 H7월전개내용(입력)
    set_widths(ws, [22, 12, 12, 12, 13, 12, 10, 46])
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    put(ws, "A1", "★ 7월 팀별 전개 체크 (부장님 피드백 반영)", H1_FONT, LEFT, NAVY)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:H4")
    guide = ("· 부장님 피드백 : 팀별 하반기 목표는 좋으나 '7월에 몇 개 했는지' 체크 + 항목별 7월 전개 내용 기입 필요\n"
             "· ① 연장진열대 : 취합 리스트 설치일자 기준 재집계 → 상반기(~6월) 75점 + 7월 신규 4점.  기존 '상반기 79점'에 섞여있던 7월 설치 4점을 분리했습니다.\n"
             "· ② 점두진열 : '진열점입력'(7월 분석) 기준 팀별 진열점 현황을 비고에 표기(날짜 없어 7월 신규분 분리 불가).\n"
             "· 노란색 칸(7월 실적 / 7월 전개 내용)은 수정 가능하며, 7월 누계·달성률(=7월 실적÷하반기 목표)은 자동 계산됩니다.")
    put(ws, "A2", guide, Font(name="맑은 고딕", size=9, color="595959"), LEFT, GREY, border=False)

    headers = ["전략 / 항목", "팀", "상반기 완료\n(~6월)", "하반기 목표",
               "7월 실적", "7월 누계", "달성률", "7월 전개 내용 / 비고"]
    hr = 6
    for i, h in enumerate(headers):
        put(ws, f"{get_column_letter(1 + i)}{hr}", h, HEAD_FONT, CENTER, BLUE)
    ws.row_dimensions[hr].height = 30

    # 각 항목: (이름, {팀:(상반기완료, 하반기목표, 7월실적, 비고)}, 지역부(상반기완료,하반기목표,비고))
    #   · 7월실적 = 숫자면 자동 기입(수정 가능), None 이면 빈 입력란
    #   · 상반기완료 = None 이면 '진행점(목표)'만 있는 항목
    items = [
        ("① 연장진열대",
         {"수원영업1팀": (14, 3, 0, ""),
          "수원영업2팀": (8, 7, 1, "권선아이파크점 7/2(진열대 전체교체)"),
          "수원영업3팀": (3, 1, 1, "화성봉황점 7/6"),
          "수원영업4팀": (6, 3, 0, ""),
          "수원영업5팀": (5, 1, 1, "산본중앙로점 7/6(양면 연장)"),
          "수원영업6팀": (9, 0, 0, ""),
          "수원영업7팀": (11, 3, 0, ""),
          "수원영업8팀": (11, 3, 0, ""),
          "수원영업9팀": (8, 5, 1, "행정드림점 7/30(전체교체·예정)")},
         (75, 26, "설치일자 기준 재집계 : ~6월 75 + 7월 4")),
        ("② 점두진열",
         {"수원영업1팀": (12, 2, None, "7월 진열점 현황 12점"),
          "수원영업2팀": (20, 3, None, "7월 진열점 현황 18점"),
          "수원영업3팀": (6, 2, None, "7월 진열점 현황 6점"),
          "수원영업4팀": (13, 8, None, "7월 진열점 현황 13점"),
          "수원영업5팀": (14, 12, None, "7월 진열점 현황 14점"),
          "수원영업6팀": (9, 11, None, "7월 진열점 현황 9점"),
          "수원영업7팀": (16, 2, None, "7월 진열점 현황 16점"),
          "수원영업8팀": (6, 4, None, "7월 진열점 현황 6점"),
          "수원영업9팀": (12, 1, None, "7월 진열점 현황 11점")},
         (108, 45, "7월 진열점 현황 105점(진열점입력 기준) · 신규분 분리 불가 → 7월 실적 직접 입력")),
        ("③ 주먹밥+음료(진행점)",
         {"수원영업1팀": (None, 14, None, ""), "수원영업2팀": (None, 13, None, ""), "수원영업3팀": (None, 9, None, ""),
          "수원영업4팀": (None, 10, None, ""), "수원영업5팀": (None, 12, None, ""), "수원영업6팀": (None, 13, None, ""),
          "수원영업7팀": (None, 6, None, ""), "수원영업8팀": (None, 1, None, ""), "수원영업9팀": (None, 9, None, "")},
         (None, 87, "제출 자료에 7월 데이터 없음 → 직접 입력")),
        ("④ 헬시플레저 샌드위치(진행점)",
         {"수원영업1팀": (None, 14, None, ""), "수원영업2팀": (None, 18, None, ""), "수원영업3팀": (None, 6, None, ""),
          "수원영업4팀": (None, 10, None, ""), "수원영업5팀": (None, 12, None, ""), "수원영업6팀": (None, 10, None, ""),
          "수원영업7팀": (None, 6, None, ""), "수원영업8팀": (None, 9, None, ""), "수원영업9팀": (None, 12, None, "")},
         (None, 97, "제출 자료에 7월 데이터 없음 → 직접 입력")),
        ("⑤ 간편식사 아크릴 진열대(진행점)",
         {"수원영업1팀": (None, 18, None, ""), "수원영업2팀": (None, 22, None, ""), "수원영업3팀": (None, 25, None, ""),
          "수원영업4팀": (None, 14, None, ""), "수원영업5팀": (None, 13, None, ""), "수원영업6팀": (None, 13, None, ""),
          "수원영업7팀": (None, 13, None, ""), "수원영업8팀": (None, 16, None, ""), "수원영업9팀": (None, 13, None, "")},
         (None, 147, "제출 자료에 7월 데이터 없음 → 직접 입력")),
        ("⑥ 라심발리 교체",
         {"수원영업1팀": (5, 1, None, ""), "수원영업2팀": (8, 1, None, ""), "수원영업3팀": (6, 4, None, ""),
          "수원영업4팀": (5, 0, None, ""), "수원영업5팀": (2, 0, None, ""), "수원영업6팀": (5, 1, None, ""),
          "수원영업7팀": (10, 2, None, ""), "수원영업8팀": (10, 2, None, ""), "수원영업9팀": (8, 1, None, "")},
         (59, 12, "제출 자료에 7월 데이터 없음 → 직접 입력")),
    ]

    r = hr + 1
    for name, teams, region in items:
        # 지역부 요약행
        c_done, c_tgt, c_memo = region
        put(ws, f"A{r}", name, BOLD, LEFT, LIGHT)
        put(ws, f"B{r}", "수원지역부(계)", BOLD, CENTER, LIGHT)
        put(ws, f"C{r}", c_done if c_done is not None else "-", BOLD, CENTER, LIGHT)
        put(ws, f"D{r}", c_tgt, BOLD, CENTER, LIGHT)
        put(ws, f"E{r}", f"=SUM(E{r+1}:E{r+len(teams)})", BOLD, CENTER, LIGHT)
        put(ws, f"F{r}", f"=SUM(F{r+1}:F{r+len(teams)})", BOLD, CENTER, LIGHT)
        put(ws, f"G{r}", f'=IFERROR(E{r}/D{r},"")', BOLD, CENTER, LIGHT, numfmt="0.0%")
        put(ws, f"H{r}", c_memo, Font(name="맑은 고딕", size=9, bold=True, color=NAVY), LEFT, LIGHT)
        r += 1
        for team, (done, tgt, jul, memo) in teams.items():
            put(ws, f"A{r}", "", BODY, LEFT)
            put(ws, f"B{r}", team, BODY, CENTER, GREY)
            put(ws, f"C{r}", done if done is not None else "-", BODY, CENTER)
            put(ws, f"D{r}", tgt, BODY, CENTER)
            # 7월 실적 : 값이 있으면 자동 기입(수정 가능), 없으면 빈 입력란
            put(ws, f"E{r}", jul, BODY, CENTER, YELLOW)
            # 7월 누계 = 상반기완료 + 7월실적 (상반기 없으면 7월실적만)
            if done is not None:
                put(ws, f"F{r}", f"=C{r}+E{r}", BODY, CENTER, numfmt="#,##0")
            else:
                put(ws, f"F{r}", f"=E{r}", BODY, CENTER, numfmt="#,##0")
            # 달성률 = 7월실적 / 하반기목표
            put(ws, f"G{r}", f'=IFERROR(E{r}/D{r},"")', BODY, CENTER, numfmt="0.0%")
            # 7월 전개 내용 / 비고 (연장·점두는 자동 기입, 그 외 입력란)
            put(ws, f"H{r}", memo if memo else None, BODY, LEFT,
                None if memo else YELLOW)
            r += 1
        r += 1

    ws.freeze_panes = "A7"


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else "수원지역부_회의자료_2026H2.xlsx"
    wb = Workbook()
    build_cover(wb)
    build_s1(wb)
    build_s2(wb)
    build_s3(wb)
    build_s4(wb)
    build_s5(wb)
    build_check(wb)
    wb.save(out)
    print("생성 완료 :", out)
    print("시트 :", wb.sheetnames)


if __name__ == "__main__":
    main()
