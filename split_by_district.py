#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
상권별 엑셀 분리 스크립트 (Python / openpyxl 사용)

원본 엑셀의 C열(상권구분) 값에 따라 개별 엑셀 파일을 생성합니다.
1행은 머리글(제목행)으로 간주하여 모든 결과 파일에 포함됩니다.

사용법:
    python split_by_district.py [원본엑셀.xlsx]
    (인자를 생략하면 같은 폴더의 .xlsx 파일을 자동으로 찾습니다.)

필요 패키지:
    pip install openpyxl
"""
import os
import re
import sys
from collections import OrderedDict

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    print('[오류] openpyxl 이 필요합니다.  설치:  pip install openpyxl')
    sys.exit(1)

KEY_COL = 3      # 상권구분 열 (A=1, B=2, C=3)
HEADER_ROW = 1   # 머리글 행 번호


def find_input(here):
    cands = [
        f for f in os.listdir(here)
        if f.lower().endswith('.xlsx')
        and not f.startswith('상권_')
        and not f.startswith('~$')
    ]
    return os.path.join(here, cands[0]) if cands else None


def sanitize(name):
    return re.sub(r'[\\/:*?"<>|]', '_', str(name)).strip()


def main():
    here = os.path.dirname(os.path.abspath(__file__))

    if len(sys.argv) > 1:
        input_path = sys.argv[1]
    else:
        input_path = find_input(here)
        if not input_path:
            print('[오류] 처리할 .xlsx 파일을 찾을 수 없습니다.')
            sys.exit(1)

    if not os.path.exists(input_path):
        print('[오류] 파일을 찾을 수 없습니다: ' + input_path)
        sys.exit(1)

    input_path = os.path.abspath(input_path)
    src_dir = os.path.dirname(input_path)
    out_dir = os.path.join(src_dir, '상권별_결과')
    os.makedirs(out_dir, exist_ok=True)

    print('=====================================================')
    print('  상권별 엑셀 생성기 (Python)')
    print('=====================================================')
    print('원본 파일 : ' + input_path)
    print('결과 폴더 : ' + out_dir)
    print()

    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) <= HEADER_ROW:
        print('[오류] 데이터 행이 없습니다.')
        sys.exit(1)

    header = rows[HEADER_ROW - 1]
    groups = OrderedDict()
    for row in rows[HEADER_ROW:]:
        key = row[KEY_COL - 1]
        if key is None or str(key).strip() == '':
            key = '(미분류)'
        key = str(key)
        groups.setdefault(key, []).append(row)

    print('상권 종류 : %d개' % len(groups))
    print('-----------------------------------------------------')

    made = 0
    for key, data_rows in groups.items():
        out_wb = openpyxl.Workbook()
        out_ws = out_wb.active
        out_ws.append(list(header))
        for row in data_rows:
            out_ws.append(list(row))
        # 머리글 굵게
        for cell in out_ws[1]:
            cell.font = Font(bold=True)

        out_path = os.path.join(out_dir, '상권_' + sanitize(key) + '.xlsx')
        out_wb.save(out_path)
        made += 1
        print('  생성됨 : %s   (%d행)' % (os.path.basename(out_path), len(data_rows)))

    print('-----------------------------------------------------')
    print('완료! 총 %d개의 엑셀 파일을 생성했습니다.' % made)
    print('결과 위치: ' + out_dir)


if __name__ == '__main__':
    main()
